import shutil
from abc import ABC, abstractmethod
from collections import OrderedDict
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Union

import openpyxl
import yaml


class Folder:

    def __init__(self, path: Path):
        self.path = path

    @property
    def name(self):
        return self.path.name

    def clear(self):
        for child in self.path.iterdir():
            if child.is_dir():
                shutil.rmtree(child)
            else:
                child.unlink()

    def subdirs(self) -> list[Path]:
        return [x for x in self.path.iterdir() if x.is_dir()]

    def files(self) -> list[Path]:
        return [x for x in self.path.iterdir() if x.is_file()]

    def contains_filename(self, filename: str) -> bool:
        for filepath in self.files():
            if filepath.name == filename:
                return True
        return False

    @contextmanager
    def clear_after(self):
        try:
            yield
        finally:
            self.clear()

    def find_by_suffix(self, suffix: str) -> list[Path]:
        return [
            child for child in self.path.iterdir()
            if child.suffix == suffix
        ]

    def find_by_name(self, name: str) -> list[Path]:
        return [
            child for child in self.path.iterdir()
            if child.name == name
        ]


class File(ABC):

    def __init__(self, path: Path) -> None:
        self._path = path

    @property
    def path(self) -> Path:
        return self._path

    def exists(self) -> bool:
        return self._path.exists()

    @abstractmethod
    def read(self) -> Any:
        ...

    @abstractmethod
    def write(self, data: Any) -> None:
        ...


class YAMLFile(File):

    def __init__(self, path: Path, default_data: Union[dict, list] = None) -> None:
        super(YAMLFile, self).__init__(path)
        if path.suffix != ".yml":
            raise Exception("The file extension must be .yml")
        if default_data and not self.exists():
            self.write(default_data)

    def read(self) -> Union[dict, list, None]:
        with open(self._path, "r", encoding="utf-8") as file:
            return yaml.safe_load(file)

    def write(self, data: Union[dict, list, None]) -> None:
        with open(self._path, "w", encoding="utf-8") as file:
            yaml.safe_dump(data, file)


class XLSXFile(File):

    def __init__(self, path: Path):
        super(XLSXFile, self).__init__(path)
        if path.suffix != ".xlsx":
            raise Exception("The file extension must be .xlsx")

    def read(self) -> list[OrderedDict]:

        wb = openpyxl.load_workbook(str(self._path))
        ws = wb.active
        if ws.max_row == 0:
            return []

        keys = []
        for column in range(1, ws.max_column + 1):
            keys.append(ws.cell(1, column).value)

        data = []
        for row in range(2, ws.max_row + 1):
            dct = OrderedDict()
            for column, key in enumerate(keys, start=1):
                dct[key] = ws.cell(row, column).value
            data.append(dct)

        return data

    def write(self, data: list[OrderedDict]) -> None:

        wb = openpyxl.Workbook()
        ws = wb.active
        if not data:
            wb.save(str(self._path))
            return

        for column, key in enumerate(data[0].keys(), start=1):
            ws.cell(1, column, key)

        for row, dct in enumerate(data, start=2):
            for column, value in enumerate(dct.values(), start=1):
                ws.cell(row, column, value)

        self._adjust_columns_width(ws)
        wb.save(str(self._path))

    @staticmethod
    def _adjust_columns_width(sheet) -> None:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # adjusted_width = (max_length + 2) * 1.2
            adjusted_width = max_length
            sheet.column_dimensions[column].width = adjusted_width
